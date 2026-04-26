"""
Attendance service — shared async logic for advanced attendance management.
Provides holiday/leave auto-marking, bulk save, export generation, and stats computation.
Also provides attendance aggregation for salary calculation integration.
"""
import calendar
import csv
import io
import logging
from datetime import date, datetime, timedelta, time as dt_time
from typing import Optional, List, Dict
import uuid

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func

from app.models.attendance_daily import AttendanceDaily, AttendanceStatus, LateMarkType
from app.models.attendance import Attendance, AttendanceType
from app.models.employee import Employee, EmployeeStatus
from app.models.holiday import Holiday
from app.models.leave import Leave, LeaveStatus, LeaveBalance
from app.services.payroll_service import (
    calculate_late_mark_type,
    calculate_working_days,
    is_second_or_fourth_saturday,
)

# Set up module-level logger
logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Status abbreviations for export
# ---------------------------------------------------------------------------
STATUS_ABBREV = {
    AttendanceStatus.PRESENT: "P",
    AttendanceStatus.ABSENT: "A",
    AttendanceStatus.HALFDAY: "H",
    AttendanceStatus.LEAVE: "L",
    AttendanceStatus.HOLIDAY: "Ho",
    AttendanceStatus.WEEKLYOFF: "WO",
}

# Excel fill colors per status
XLSX_FILLS = {
    AttendanceStatus.PRESENT: "C6EFCE",
    AttendanceStatus.ABSENT: "FFC7CE",
    AttendanceStatus.HALFDAY: "E1D5E7",
    AttendanceStatus.LEAVE: "FFE0B2",
    AttendanceStatus.HOLIDAY: "DDEEFF",
    AttendanceStatus.WEEKLYOFF: "F2F2F2",
}


# ---------------------------------------------------------------------------
# Helper: resolve a single day's status given context
# ---------------------------------------------------------------------------
def _resolve_day(
    d: date,
    holiday_dates: set,
    leave_dates: set,
    existing: Optional[AttendanceDaily],
) -> dict:
    """
    Apply priority rules for a single calendar day and return a resolved record dict.

    Priority:
      1. is_overridden=True on existing record → preserve as-is
      2. Sunday or 2nd/4th Saturday → weeklyoff
      3. date in holiday_dates → holiday
      4. date in leave_dates → leave
      5. existing record exists → use it
      6. default → not marked (status=None)
    """
    # 1. Override preservation
    if existing is not None and existing.is_overridden:
        return {
            "id": existing.id,
            "status": existing.status,
            "check_in": existing.check_in,
            "check_out": existing.check_out,
            "late_mark_type": existing.late_mark_type,
            "is_overridden": True,
        }

    # 2. Weekly off
    if d.weekday() == 6 or is_second_or_fourth_saturday(d):
        return {
            "id": existing.id if existing else None,
            "status": AttendanceStatus.WEEKLYOFF,
            "check_in": None,
            "check_out": None,
            "late_mark_type": LateMarkType.NONE,
            "is_overridden": False,
        }

    # 3. Holiday
    if d in holiday_dates:
        return {
            "id": existing.id if existing else None,
            "status": AttendanceStatus.HOLIDAY,
            "check_in": None,
            "check_out": None,
            "late_mark_type": LateMarkType.NONE,
            "is_overridden": False,
        }

    # 4. Leave
    if d in leave_dates:
        return {
            "id": existing.id if existing else None,
            "status": AttendanceStatus.LEAVE,
            "check_in": None,
            "check_out": None,
            "late_mark_type": LateMarkType.NONE,
            "is_overridden": False,
        }

    # 5. Existing record
    if existing is not None:
        return {
            "id": existing.id,
            "status": existing.status,
            "check_in": existing.check_in,
            "check_out": existing.check_out,
            "late_mark_type": existing.late_mark_type,
            "is_overridden": existing.is_overridden,
        }

    # 6. Default: not marked (no record exists)
    return {
        "id": None,
        "status": None,
        "check_in": None,
        "check_out": None,
        "late_mark_type": LateMarkType.NONE,
        "is_overridden": False,
    }


# ---------------------------------------------------------------------------
# 1. apply_holiday_leave_automark
# ---------------------------------------------------------------------------
async def apply_holiday_leave_automark(
    db: AsyncSession,
    emp_id: str,
    month: int,
    year: int,
    records: dict,  # dict[date, AttendanceDaily] — existing records keyed by date
) -> dict:  # dict[date, dict] — resolved records
    """
    For each calendar day in the month, apply the priority resolution:
      1. is_overridden=True → preserve
      2. Weekly off (Sunday / 2nd-4th Saturday) → weeklyoff
      3. Active holiday → holiday
      4. Approved leave → leave
      5. Existing record → use it
      6. Default → absent

    Returns a dict keyed by date with resolved record dicts.
    """
    _, days_in_month = calendar.monthrange(year, month)
    first_day = date(year, month, 1)
    last_day = date(year, month, days_in_month)

    # Fetch active holidays for the month
    holiday_result = await db.execute(
        select(Holiday).where(
            and_(
                Holiday.year == year,
                Holiday.is_active == True,  # noqa: E712
                func.strftime("%m", Holiday.date) == f"{month:02d}",
            )
        )
    )
    holidays = holiday_result.scalars().all()
    holiday_dates: set = {h.date for h in holidays}

    # Fetch approved leaves for the employee that overlap with the month
    leave_result = await db.execute(
        select(Leave).where(
            and_(
                Leave.emp_id == emp_id,
                Leave.status == LeaveStatus.APPROVED,
                Leave.from_date <= last_day,
                Leave.to_date >= first_day,
            )
        )
    )
    leaves = leave_result.scalars().all()

    # Expand leave ranges to individual dates within the month
    leave_dates: set = set()
    for leave in leaves:
        current = max(leave.from_date, first_day)
        end = min(leave.to_date, last_day)
        while current <= end:
            leave_dates.add(current)
            current += timedelta(days=1)

    # Resolve each calendar day
    resolved: dict = {}
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        existing = records.get(d)
        resolved[d] = _resolve_day(d, holiday_dates, leave_dates, existing)

    return resolved


# ---------------------------------------------------------------------------
# 2. bulk_save_records
# ---------------------------------------------------------------------------
async def bulk_save_records(
    db: AsyncSession,
    emp_id: str,
    month: int,
    year: int,
    records: list,  # list of dicts with {date, status, check_in?, check_out?}
    override_by: str,
) -> dict:  # {created: N, updated: M, total: N+M}
    """
    Upsert all provided records into attendance_daily within a single transaction.
    Calculates late_mark_type for present/halfday records with check_in.
    Returns {created, updated, total}.
    Rolls back the entire transaction on any failure.
    """
    created = 0
    updated = 0

    try:
        for record in records:
            record_date = record["date"]
            status = record["status"]
            check_in_str = record.get("check_in")
            check_out_str = record.get("check_out")

            # Parse check_in / check_out from "HH:MM" string to datetime
            check_in_dt: Optional[datetime] = None
            check_out_dt: Optional[datetime] = None

            if check_in_str:
                try:
                    t = datetime.strptime(check_in_str, "%H:%M").time()
                    check_in_dt = datetime(
                        record_date.year, record_date.month, record_date.day,
                        t.hour, t.minute
                    )
                except (ValueError, AttributeError):
                    check_in_dt = None

            if check_out_str:
                try:
                    t = datetime.strptime(check_out_str, "%H:%M").time()
                    check_out_dt = datetime(
                        record_date.year, record_date.month, record_date.day,
                        t.hour, t.minute
                    )
                except (ValueError, AttributeError):
                    check_out_dt = None

            # Calculate late_mark_type
            late_mark_type = LateMarkType.NONE
            if check_in_dt and status in (
                AttendanceStatus.PRESENT, AttendanceStatus.HALFDAY,
                "present", "halfday",
            ):
                late_mark_type = calculate_late_mark_type(check_in_dt.time())

            # Determine is_late_mark / is_half_late_mark / is_half_day flags
            is_late_mark = late_mark_type == LateMarkType.LATE
            is_half_late_mark = late_mark_type == LateMarkType.HALF_LATE
            is_half_day = (
                status in (AttendanceStatus.HALFDAY, "halfday")
                or late_mark_type == LateMarkType.HALF_DAY
            )

            # Check for existing record
            existing_result = await db.execute(
                select(AttendanceDaily).where(
                    and_(
                        AttendanceDaily.emp_id == emp_id,
                        AttendanceDaily.date == record_date,
                    )
                )
            )
            existing = existing_result.scalar_one_or_none()

            if existing is not None:
                # UPDATE
                existing.status = status
                existing.check_in = check_in_dt
                existing.check_out = check_out_dt
                existing.late_mark_type = late_mark_type
                existing.is_late_mark = is_late_mark
                existing.is_half_late_mark = is_half_late_mark
                existing.is_half_day = is_half_day
                existing.is_overridden = True
                existing.override_by = override_by
                existing.updated_at = datetime.utcnow()
                db.add(existing)
                updated += 1
            else:
                # INSERT
                new_record = AttendanceDaily(
                    id=str(uuid.uuid4()),
                    emp_id=emp_id,
                    date=record_date,
                    check_in=check_in_dt,
                    check_out=check_out_dt,
                    status=status,
                    late_mark_type=late_mark_type,
                    is_late_mark=is_late_mark,
                    is_half_late_mark=is_half_late_mark,
                    is_half_day=is_half_day,
                    is_overridden=True,
                    override_by=override_by,
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow(),
                )
                db.add(new_record)
                created += 1

        await db.commit()

    except Exception:
        await db.rollback()
        raise

    return {"created": created, "updated": updated, "total": created + updated}


# ---------------------------------------------------------------------------
# 3. generate_export
# ---------------------------------------------------------------------------
async def generate_export(
    db: AsyncSession,
    month: int,
    year: int,
    format: str,  # "csv" or "xlsx"
    emp_id: Optional[str] = None,
) -> tuple:  # (bytes, filename)
    """
    Build an employee × day attendance matrix and return as (bytes, filename).
    Applies holiday/leave auto-marking before building the matrix.
    """
    _, days_in_month = calendar.monthrange(year, month)

    # Fetch employees
    if emp_id:
        emp_result = await db.execute(
            select(Employee).where(Employee.id == emp_id)
        )
        employees = emp_result.scalars().all()
    else:
        emp_result = await db.execute(
            select(Employee).where(Employee.status == EmployeeStatus.ACTIVE)
        )
        employees = emp_result.scalars().all()

    # Fetch all attendance records for the month (for all relevant employees)
    emp_ids = [e.id for e in employees]
    first_day = date(year, month, 1)
    last_day = date(year, month, days_in_month)

    if emp_ids:
        att_result = await db.execute(
            select(AttendanceDaily).where(
                and_(
                    AttendanceDaily.emp_id.in_(emp_ids),
                    AttendanceDaily.date >= first_day,
                    AttendanceDaily.date <= last_day,
                )
            )
        )
        all_records = att_result.scalars().all()
    else:
        all_records = []

    # Group records by emp_id → date
    records_by_emp: Dict[str, Dict[date, AttendanceDaily]] = {}
    for rec in all_records:
        records_by_emp.setdefault(rec.emp_id, {})[rec.date] = rec

    # Fetch active holidays for working_days calculation
    holiday_result = await db.execute(
        select(Holiday).where(
            and_(
                Holiday.year == year,
                Holiday.is_active == True,  # noqa: E712
                func.strftime("%m", Holiday.date) == f"{month:02d}",
            )
        )
    )
    holidays = holiday_result.scalars().all()
    holiday_date_list = [h.date for h in holidays]
    working_days = calculate_working_days(year, month, holiday_date_list)

    # Build header
    day_headers = [str(d) for d in range(1, days_in_month + 1)]
    header = (
        ["Emp Code", "Name", "Department"]
        + day_headers
        + ["Working Days", "Present", "Absent", "Late Marks", "Half Late", "LOP"]
    )

    # Build rows
    rows = []
    for emp in employees:
        emp_records = records_by_emp.get(emp.id, {})
        resolved = await apply_holiday_leave_automark(
            db, emp.id, month, year, emp_records
        )

        # Day cells
        day_cells = []
        present_count = 0
        absent_count = 0
        late_mark_count = 0
        half_late_count = 0
        lop_days = 0

        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            rec = resolved.get(d)
            if rec is None:
                day_cells.append("-")
                continue

            status = rec["status"]
            abbrev = STATUS_ABBREV.get(status, "-")
            day_cells.append(abbrev)

            if status == AttendanceStatus.PRESENT:
                present_count += 1
            elif status == AttendanceStatus.ABSENT:
                absent_count += 1
                lop_days += 1
            elif status == AttendanceStatus.HALFDAY:
                present_count += 1  # halfday counts as partial present
                lop_days += 0.5

            # Late mark counts
            late_mark = rec.get("late_mark_type", LateMarkType.NONE)
            if late_mark == LateMarkType.LATE:
                late_mark_count += 1
            elif late_mark == LateMarkType.HALF_LATE:
                half_late_count += 1

        row = (
            [emp.emp_code, emp.name, emp.department or ""]
            + day_cells
            + [
                working_days,
                present_count,
                absent_count,
                late_mark_count,
                half_late_count,
                lop_days,
            ]
        )
        rows.append(row)

    # Generate output
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(header)
        writer.writerows(rows)
        file_bytes = output.getvalue().encode("utf-8")
        filename = f"attendance_{month}_{year}.csv"
        return file_bytes, filename

    elif format == "xlsx":
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("Excel export unavailable: openpyxl is not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {month}-{year}"

        # Write header row with bold font
        ws.append(header)
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        # Status → fill color mapping (by abbreviation)
        abbrev_to_fill = {
            abbrev: PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            for status, color in XLSX_FILLS.items()
            for abbrev in [STATUS_ABBREV.get(status, "")]
            if abbrev
        }

        # Write data rows with color coding
        for row_data in rows:
            ws.append(row_data)
            row_idx = ws.max_row
            # Color day cells (columns 4 to 3+days_in_month)
            for col_offset, day in enumerate(range(1, days_in_month + 1)):
                col_idx = 3 + col_offset + 1  # 1-indexed; cols 1-3 are Emp Code/Name/Dept
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                if cell_value and cell_value in abbrev_to_fill:
                    cell.fill = abbrev_to_fill[cell_value]

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        file_bytes = output.getvalue()
        filename = f"attendance_{month}_{year}.xlsx"
        return file_bytes, filename

    else:
        raise ValueError(f"Unsupported format: {format}. Must be 'csv' or 'xlsx'.")


# ---------------------------------------------------------------------------
# 4. compute_stats
# ---------------------------------------------------------------------------
async def compute_stats(
    db: AsyncSession,
    month: int,
    year: int,
) -> dict:
    """
    Compute attendance statistics for the given month/year.
    Returns:
      {
        summary_cards: {working_days, present_pct, total_late_marks, total_lop_days},
        daily_trend: [{day, present_count, date}],
        department_stats: [{department, present, absent, late_mark, present_pct}],
      }
    """
    _, days_in_month = calendar.monthrange(year, month)
    first_day = date(year, month, 1)
    last_day = date(year, month, days_in_month)

    # Fetch all active employees
    emp_result = await db.execute(
        select(Employee).where(Employee.status == EmployeeStatus.ACTIVE)
    )
    employees = emp_result.scalars().all()
    total_employees = len(employees)

    # Fetch active holidays for working_days calculation
    holiday_result = await db.execute(
        select(Holiday).where(
            and_(
                Holiday.year == year,
                Holiday.is_active == True,  # noqa: E712
                func.strftime("%m", Holiday.date) == f"{month:02d}",
            )
        )
    )
    holidays = holiday_result.scalars().all()
    holiday_date_list = [h.date for h in holidays]
    working_days = calculate_working_days(year, month, holiday_date_list)

    # Fetch all attendance_daily records for the month (with employee join for department)
    att_result = await db.execute(
        select(AttendanceDaily, Employee.department).join(
            Employee, AttendanceDaily.emp_id == Employee.id
        ).where(
            and_(
                AttendanceDaily.date >= first_day,
                AttendanceDaily.date <= last_day,
            )
        )
    )
    att_rows = att_result.all()

    # Early return when no data
    if not att_rows and total_employees == 0:
        return {
            "summary_cards": {
                "working_days": working_days,
                "present_pct": 0.0,
                "total_late_marks": 0,
                "total_lop_days": 0.0,
            },
            "daily_trend": [
                {
                    "day": d,
                    "present_count": 0,
                    "date": str(date(year, month, d)),
                }
                for d in range(1, days_in_month + 1)
            ],
            "department_stats": [],
        }

    # Build daily_trend: count present/halfday per day
    daily_present: Dict[int, int] = {d: 0 for d in range(1, days_in_month + 1)}
    for att, _dept in att_rows:
        if att.status in (AttendanceStatus.PRESENT, AttendanceStatus.HALFDAY):
            daily_present[att.date.day] = daily_present.get(att.date.day, 0) + 1

    daily_trend = [
        {
            "day": d,
            "present_count": daily_present.get(d, 0),
            "date": str(date(year, month, d)),
        }
        for d in range(1, days_in_month + 1)
    ]

    # Build department_stats
    dept_data: Dict[str, Dict] = {}
    for att, dept in att_rows:
        dept_key = dept or "Unknown"
        if dept_key not in dept_data:
            dept_data[dept_key] = {"present": 0, "absent": 0, "late_mark": 0, "total": 0}
        dept_data[dept_key]["total"] += 1
        if att.status in (AttendanceStatus.PRESENT, AttendanceStatus.HALFDAY):
            dept_data[dept_key]["present"] += 1
        elif att.status == AttendanceStatus.ABSENT:
            dept_data[dept_key]["absent"] += 1
        if att.is_late_mark:
            dept_data[dept_key]["late_mark"] += 1

    department_stats = []
    for dept_name, data in dept_data.items():
        total = data["total"]
        present = data["present"]
        present_pct = round((present / total) * 100, 2) if total > 0 else 0.0
        department_stats.append(
            {
                "department": dept_name,
                "present": present,
                "absent": data["absent"],
                "late_mark": data["late_mark"],
                "present_pct": present_pct,
            }
        )

    # Summary cards
    total_present_records = sum(
        1 for att, _ in att_rows
        if att.status in (AttendanceStatus.PRESENT, AttendanceStatus.HALFDAY)
    )
    denominator = total_employees * working_days
    present_pct = round((total_present_records / denominator) * 100, 2) if denominator > 0 else 0.0

    total_late_marks = sum(1 for att, _ in att_rows if att.is_late_mark)

    # Total LOP days: sum lwp_days from leave_balances for the year
    lop_result = await db.execute(
        select(func.sum(LeaveBalance.lwp_days)).where(
            LeaveBalance.year == year
        )
    )
    total_lop_days = float(lop_result.scalar() or 0)

    return {
        "summary_cards": {
            "working_days": working_days,
            "present_pct": present_pct,
            "total_late_marks": total_late_marks,
            "total_lop_days": total_lop_days,
        },
        "daily_trend": daily_trend,
        "department_stats": department_stats,
    }


# ---------------------------------------------------------------------------
# Salary Calculation Integration Functions
# ---------------------------------------------------------------------------

def _count_working_days(start_date: date, end_date: date) -> int:
    """
    Count weekdays (Monday-Friday) between start_date and end_date (inclusive).
    Excludes weekends (Saturday=5, Sunday=6 in weekday()).
    
    Args:
        start_date: Start date of the period (inclusive)
        end_date: End date of the period (inclusive)
    
    Returns:
        Integer count of working days (weekdays only)
    
    Examples:
        >>> _count_working_days(date(2024, 1, 1), date(2024, 1, 7))  # Mon-Sun
        5
        >>> _count_working_days(date(2024, 1, 1), date(2024, 1, 5))  # Mon-Fri
        5
    """
    if start_date > end_date:
        return 0
    
    working_days = 0
    current = start_date
    
    while current <= end_date:
        # weekday(): Monday=0, Sunday=6
        # Exclude Saturday(5) and Sunday(6)
        if current.weekday() < 5:  # Monday(0) to Friday(4)
            working_days += 1
        current += timedelta(days=1)
    
    return working_days


def _calculate_present_days(attendance_records: List[Attendance]) -> int:
    """
    Count unique dates with at least one check-in record.
    Multiple check-ins on the same date count as 1 present day.
    
    Args:
        attendance_records: List of Attendance records with check_in_time
    
    Returns:
        Integer count of unique present days
    
    Examples:
        If records have check-ins on [2024-01-01, 2024-01-01, 2024-01-02]:
        Returns 2 (two unique dates)
    """
    unique_dates = set()
    
    for record in attendance_records:
        # Only count records with non-null date and CHECK_IN type
        if record.date and record.attendance_type == AttendanceType.CHECK_IN:
            unique_dates.add(record.date)
    
    return len(unique_dates)


def _calculate_overtime_hours(
    attendance_records: List[Attendance], 
    standard_hours: float = 8.0
) -> float:
    """
    Calculate total overtime hours from check-in/check-out pairs.
    Overtime = max(0, working_hours - standard_hours) per day.
    
    Args:
        attendance_records: List of Attendance records
        standard_hours: Standard shift duration in hours (default: 8.0)
    
    Returns:
        Float value of total overtime hours, rounded to 2 decimal places
    
    Logic:
        - Groups records by date
        - For each date with both CHECK_IN and CHECK_OUT, calculates working hours
        - Overtime = max(0, working_hours - standard_hours)
        - Sums overtime across all days
    """
    # Group records by date
    records_by_date: Dict[date, Dict[str, Optional[datetime]]] = {}
    
    for record in attendance_records:
        if not record.date:
            continue
        
        if record.date not in records_by_date:
            records_by_date[record.date] = {"check_in": None, "check_out": None}
        
        # Combine date and time to create datetime
        if record.time:
            record_datetime = datetime.combine(record.date, record.time)
            
            if record.attendance_type == AttendanceType.CHECK_IN:
                # Keep earliest check-in
                if records_by_date[record.date]["check_in"] is None:
                    records_by_date[record.date]["check_in"] = record_datetime
                else:
                    records_by_date[record.date]["check_in"] = min(
                        records_by_date[record.date]["check_in"], 
                        record_datetime
                    )
            elif record.attendance_type == AttendanceType.CHECK_OUT:
                # Keep latest check-out
                if records_by_date[record.date]["check_out"] is None:
                    records_by_date[record.date]["check_out"] = record_datetime
                else:
                    records_by_date[record.date]["check_out"] = max(
                        records_by_date[record.date]["check_out"], 
                        record_datetime
                    )
    
    # Calculate overtime for each day
    total_overtime = 0.0
    
    for day_data in records_by_date.values():
        check_in = day_data["check_in"]
        check_out = day_data["check_out"]
        
        if check_in and check_out and check_out > check_in:
            # Calculate working hours
            time_diff = check_out - check_in
            working_hours = time_diff.total_seconds() / 3600.0  # Convert to hours
            
            # Calculate overtime (only positive values)
            daily_overtime = max(0.0, working_hours - standard_hours)
            total_overtime += daily_overtime
    
    # Round to 2 decimal places
    return round(total_overtime, 2)


async def _get_approved_leave_days(
    db: AsyncSession,
    employee_id: str,
    start_date: date,
    end_date: date
) -> int:
    """
    Query and sum approved leave days for an employee within a date range.
    
    Args:
        db: Async database session
        employee_id: Employee UUID
        start_date: Period start date
        end_date: Period end date
    
    Returns:
        Integer count of approved leave days in the period
    
    Logic:
        - Queries Leave table for APPROVED leaves
        - Filters for leaves that overlap with the date range
        - Sums total_days from matching records
    """
    try:
        result = await db.execute(
            select(func.sum(Leave.total_days)).where(
                and_(
                    Leave.emp_id == employee_id,
                    Leave.status == LeaveStatus.APPROVED,
                    Leave.from_date <= end_date,
                    Leave.to_date >= start_date
                )
            )
        )
        
        total_leave_days = result.scalar()
        
        # Handle None case (no leaves found)
        if total_leave_days is None:
            return 0
        
        # Convert Decimal to int
        return int(total_leave_days)
    
    except Exception as e:
        logger.error(f"Error fetching approved leave days for employee {employee_id}: {e}")
        return 0


async def get_employee_attendance_summary(
    db: AsyncSession,
    employee_id: str,
    start_date: date,
    end_date: date
) -> dict:
    """
    Aggregate attendance data for an employee within a date range.
    
    This function fetches attendance records, calculates present days, overtime hours,
    approved leave days, and computes absent days based on working days.
    
    Args:
        db: Async database session
        employee_id: Employee UUID
        start_date: Period start date (inclusive)
        end_date: Period end date (inclusive)
    
    Returns:
        Dictionary with attendance summary:
        {
            "present_days": int,      # Days with at least one check-in
            "absent_days": int,       # working_days - present_days - leave_days
            "leave_days": int,        # Approved leave days
            "overtime_hours": float,  # Total OT hours (rounded to 2 decimals)
            "total_days": int,        # Calendar days in period
            "working_days": int       # Weekdays (Mon-Fri) in period
        }
    
    Error Handling:
        - On database errors, returns default values with full LOP
        - Logs warnings for zero attendance cases
        - Validates data integrity and adjusts if needed
    """
    try:
        logger.info(
            f"Fetching attendance for employee {employee_id} from {start_date} to {end_date}"
        )
        
        # Query attendance_daily records for the employee in the date range
        # Use attendance_daily table which has daily status (PRESENT, ABSENT, etc.)
        result = await db.execute(
            select(AttendanceDaily).where(
                and_(
                    AttendanceDaily.emp_id == employee_id,
                    AttendanceDaily.date >= start_date,
                    AttendanceDaily.date <= end_date
                )
            ).order_by(AttendanceDaily.date)
        )
        
        daily_records = result.scalars().all()
        
        # Convert daily records to attendance records format for helper functions
        # Count present days from attendance_daily
        # NOTE: Database stores status as string (e.g., 'PRESENT'), not enum object
        # So we compare with enum.value (string) instead of enum object
        present_days = sum(1 for r in daily_records if r.status == AttendanceStatus.PRESENT.value)
        
        # Count half days from attendance_daily
        # Half days count as 0.5 present for attendance, but 0.5 LOP for salary
        halfday_count = sum(1 for r in daily_records if r.status == AttendanceStatus.HALFDAY.value)
        
        # Count leave days from attendance_daily (not from leaves table)
        # This ensures attendance_daily is the single source of truth
        leave_days = sum(1 for r in daily_records if r.status == AttendanceStatus.LEAVE.value)
        
        # Count absent days directly from attendance_daily
        # This is more reliable than calculating from formula
        absent_count_from_records = sum(1 for r in daily_records if r.status == AttendanceStatus.ABSENT.value)
        
        # Get effective policy for shift hours and OT multiplier
        from app.services.policy_service import get_policy
        from app.services.payroll_service import get_effective_policy
        from app.models.policy import EmployeePolicyOverride
        company_policy = await get_policy(db)
        override_stmt = select(EmployeePolicyOverride).where(EmployeePolicyOverride.emp_id == employee_id)
        override_result = await db.execute(override_stmt)
        override = override_result.scalars().first()
        policy = get_effective_policy(company_policy, override)
        
        shift_hours_val = float(policy.shift_hours) if policy.shift_hours else 8.0
        ot_multiplier_val = float(policy.ot_normal_multiplier) if policy.ot_normal_multiplier else 2.0

        # Calculate overtime from attendance_daily check_in/check_out times
        # Include both PRESENT and HALFDAY records for OT calculation
        overtime_hours = 0.0
        for record in daily_records:
            # Check if record has times AND status is PRESENT or HALFDAY
            if record.check_in and record.check_out and record.status in (
                AttendanceStatus.PRESENT.value, 
                AttendanceStatus.HALFDAY.value
            ):
                time_diff = record.check_out - record.check_in
                working_hours = time_diff.total_seconds() / 3600.0
                daily_overtime = max(0.0, working_hours - shift_hours_val)  # Use policy shift hours
                overtime_hours += daily_overtime
        overtime_hours = round(overtime_hours, 2)
        
        # Calculate working days (weekdays only)
        working_days = _count_working_days(start_date, end_date)
        
        # Calculate total calendar days
        total_days = (end_date - start_date).days + 1
        
        # Calculate absent days
        # Use direct count from attendance_daily instead of formula
        # This is more reliable and matches the actual attendance records
        absent_days = absent_count_from_records
        
        # Data validation and integrity checks
        if absent_days < 0:
            absent_days = 0
        
        # Validate overtime hours

        if overtime_hours < 0:
            logger.warning(
                f"Invalid overtime_hours ({overtime_hours}) for employee {employee_id}. "
                f"Setting to 0."
            )
            overtime_hours = 0.0
        
        # Log warning for zero attendance
        if len(daily_records) == 0:
            logger.warning(
                f"No attendance records found for employee {employee_id} "
                f"in period {start_date} to {end_date}"
            )
        
        return {
            "present_days": present_days,
            "absent_days": absent_days,
            "leave_days": leave_days,
            "halfday_count": halfday_count,
            "overtime_hours": overtime_hours,
            "total_days": total_days,
            "working_days": working_days,
            "shift_hours": shift_hours_val,
            "ot_multiplier": ot_multiplier_val,
        }
    
    except Exception as e:
        logger.error(
            f"Database error fetching attendance for employee {employee_id}: {e}",
            exc_info=True
        )
        
        # Return default values on error (full LOP scenario)
        working_days = _count_working_days(start_date, end_date)
        total_days = (end_date - start_date).days + 1
        
        return {
            "present_days": 0,
            "absent_days": working_days,
            "leave_days": 0,
            "overtime_hours": 0.0,
            "total_days": total_days,
            "working_days": working_days,
            "shift_hours": 8.0,
            "ot_multiplier": 2.0,
        }
